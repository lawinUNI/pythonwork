import openpyxl as op

# Excel ファイル（ワークブック）の新規作成
wb = op.Workbook()

# ワークシートの有効化
ws = wb.active

# Excel ファイルを Google ドライブに保存
# Google ドライブ上の "/content/drive/MyDrive" ディレクトリに "python-excel.xlsx" というファイルが作成される
wb.save('c:\\work\\aaa.xlsx')


ws['A1'] = 'あいうえお'

wb.save('c:\\work\\aaa.xlsx')

print(ws['A1'].value)

wb = op.load_workbook('c:\\work\\nsm8_1_tanjun.xlsx')

print(wb.sheetnames)

ws = wb['単純集計表']

print(ws['A1'].value)

for colobj in list(ws.columns)[1]:
    print(colobj.value)


